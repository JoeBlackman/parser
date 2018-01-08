import csv
import openpyxl #ignored error for unidentified reference here (error happened after importing this script to git)
import sys
import argparse
from datetime import date

#input: path to excel workbook file. output: workbook object
def getWorkbook(workbookName):
    workbook = openpyxl.load_workbook(workbookName)
    return workbook

def readTxtFile(txtFile):
    with open(txtFile, 'r') as f:
        headerList = f.read()
        f.close()
    return headerList

#input: workbook object, name of sheet in workbook object. output: sheet object
def getSheet(workbook, sheetName):
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet

def unpack(worksheet):
    data = [list(x) for x in worksheet] #list comprehension
    return data

#input: input file path, sheet name, output file path, data to be written. csv created or overwritten to output file path
def makeCSV(inputFileName, sheetName, outputFileName, table):
    with open(outputFileName, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([inputFileName]) #add a header with the path used to create this csv
        writer.writerow([sheetName]) #add header with the sheet name used to create this csv
        writer.writerow([date.today()])
        writer.writerow([''])
        for row in table: #for each row of data, write it to the csv
            writer.writerow(row)
        f.close()

#input a list of [registers:startbit-stopbit]. output list of lists [register, bit, length]
def splitRegisterContent(listOfReg):
    registerAddress = []
    startBit = []
    numOfBits = []
    registerAddress.append("Register Address")
    startBit.append("Start Bit")
    numOfBits.append("Bitwise Length")

    #need to skip first row of list
    for i in range(1, len(listOfReg)):
        val = listOfReg[i]
        registerAddress.append(val[:5])
        if ':' in val:
            if '-' in val:  # multiple bits
                startBit.append(str(val[(val.index('-') - 2):(val.index('-'))]))
                numOfBits.append(str((int(val[(val.index('-') + 1):(val.index('-') + 3)]) - int(
                    val[(val.index('-') - 2):(val.index('-'))]))))
            else:  # single bit
                startBit.append(str(val[(val.index(':') + 1):(val.index(':') + 3)]))
                numOfBits.append("1")
        else:
            startBit.append("0")
            numOfBits.append("All")
    return [registerAddress, startBit, numOfBits]

def main(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--iFile", required=True, help="Specifies the name of the workbook you wish to import for parsing")
    parser.add_argument("-s", "--sName", required=True, help="Specifies the name of the sheet in the workbook you wish to parse")
    parser.add_argument("-c", "--columns", help="Specifies the name of the text file with a list of columns the user wants in the output")
    parser.add_argument("-o", "--oName", required=True, help="Specifies the name of the output file you wish to write the parsed data to")
    args = parser.parse_args()

    inputPDSName = args.iFile
    sheetName = args.sName
    inputHeadersName = args.columns
    outputFileName = args.oName

    pds = []
    sheet = []
    try:
        pds = getWorkbook(inputPDSName)
    except FileNotFoundError as err:
        exit(err)
    except PermissionError as err:
        exit(err)
    except TypeError as err:
        exit(err)
    try:
        sheet = getSheet(pds, sheetName)
    except KeyError as err:
        exit(err)

    #extract values from cells, store in list of lists
    data = []
    for row in sheet.rows:
        rowValues = []
        for cell in row:
            value = cell.value
            rowValues.append(value)
        data.append(rowValues)

    originalData = [list(x) for x in data]  # list comprehension

    originalHeaders = []
    for column in originalData:
        header = column[0]
        originalHeaders.append(header)

    # default to keeping all columns if no --columns arg specified
    desiredHeaders = []
    if inputHeadersName != None:
        desiredHeaders = readTxtFile(inputHeadersName)
    else:
        desiredHeaders = originalHeaders

    #checks for headers that actually exist
    matchingHeaders = []
    for header in desiredHeaders:
        if header in originalHeaders:
            matchingHeaders.append(header)

    #if the user specified a register address column, extra parsing is required otherwise, we just return everything
    #that they ask for
    finalDataTable = []
    if 'Register Address' in matchingHeaders:
        registerColumn = []
        #search for the column containing register data an copy it
        for column in originalData:
            header = column[0]
            if header == 'Register Address':
                registerColumn = column
        #perform split of column
        splitRegister = splitRegisterContent(registerColumn)
        #add the 3 columns of the newly split register column to the final data table
        finalDataTable.append(splitRegister[0])
        finalDataTable.append(splitRegister[1])
        finalDataTable.append(splitRegister[2])
        for column in originalData:
            header = column[0]
            for name in matchingHeaders:
                if (name == header) & (header != 'Register Address'):
                    finalDataTable.append(column)
                    #if header != 'Register Address':
                        #finalDataTable.append(column)
    else:
        #no registerColumn specified. don't need to unpack a register column then, just build a list without it
        for column in originalData:
            header = column[0]
            for name in matchingHeaders:
                if name == header:
                    finalDataTable.append(column)#will only add columns to the final data table if the headers match

    finalDataTable = zip(*finalDataTable) #formats columns side by side

    try:
        makeCSV(inputPDSName, sheetName, outputFileName, finalDataTable)
    except FileNotFoundError as err:
        #print(err)
        exit(err)
    except PermissionError as err:
        #print(err)
        exit(err)
    except TypeError as err:
        #print(err)
        exit(err)

    print("OK")
    exit(0)

if __name__ == "__main__":
    main(sys.argv[1:])